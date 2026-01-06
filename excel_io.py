import os
import sqlite3
from openpyxl import Workbook, load_workbook


TABLES = [
    "settings",
    "admins",
    "admins_pending",
    "bans",
    "bans_pending",
    "system_prompts",
    "users",
    "user_changes",
    "messages",
]


def export_db_to_xlsx(db_path: str, out_path: str) -> None:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    wb = Workbook()
    wb.remove(wb.active)

    for t in TABLES:
        ws = wb.create_sheet(title=t)
        rows = conn.execute(f"SELECT * FROM {t}").fetchall()
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r[h] for h in headers])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    conn.close()


def import_xlsx_to_db(xlsx_path: str, db_path: str) -> None:
    """
    Replaces table contents from sheets.
    SECURITY: only run this for trusted admin imports.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        cur.execute("BEGIN;")
        for t in TABLES:
            if t not in wb.sheetnames:
                continue
            ws = wb[t]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) for h in rows[0]]
            data_rows = rows[1:]

            cur.execute(f"DELETE FROM {t};")
            if not data_rows:
                continue

            placeholders = ",".join(["?"] * len(headers))
            cols = ",".join(headers)
            sql = f"INSERT INTO {t} ({cols}) VALUES ({placeholders});"
            for dr in data_rows:
                cur.execute(sql, list(dr))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
